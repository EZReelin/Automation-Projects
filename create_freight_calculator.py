#!/usr/bin/env python3
"""
Die Co. Freight Calculator Excel Workbook Generator
Creates an Excel workbook for calculating and allocating freight costs
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

def create_freight_calculator():
    """Create the Die Co. Freight Calculator workbook"""

    wb = Workbook()

    # Remove default sheet and create our sheets
    wb.remove(wb.active)
    ws1 = wb.create_sheet("Import Data Entry", 0)
    ws2 = wb.create_sheet("Freight Calculator", 1)
    ws3 = wb.create_sheet("ERP Upload", 2)

    # Define styles
    header_font = Font(bold=True, size=14)
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_font = Font(bold=True, color="FF0000")
    center_align = Alignment(horizontal="center", vertical="center")
    currency_format = '"$"#,##0.00'
    percentage_format = '0.00%'
    date_format = 'MM/DD/YYYY'
    number_format = '#,##0'

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ============================================================================
    # SHEET 1: Import Data Entry
    # ============================================================================

    # Section A: Entry Summary (Rows 1-18)
    ws1.merge_cells('A1:F1')
    ws1['A1'] = "DIE CO., INC. - IMPORT FREIGHT CALCULATOR"
    ws1['A1'].font = header_font
    ws1['A1'].alignment = center_align

    ws1['A2'] = "Entry Summary Data"
    ws1['A2'].font = bold_font

    # Entry information
    ws1['A3'] = "Entry #:"
    ws1['A4'] = "Entry Date:"
    ws1['A5'] = "Vessel:"
    ws1['A6'] = "Invoice #:"

    # Format date cell
    ws1['B4'].number_format = date_format

    # Cost breakdown
    ws1['A8'] = "Total Duty (Box 37):"
    ws1['C8'] = "From CBP Entry Summary"
    ws1['B8'].number_format = currency_format

    ws1['A9'] = "Customs Entry Fee:"
    ws1['B9'].number_format = currency_format

    ws1['A10'] = "ISF - Security Filing:"
    ws1['B10'].number_format = currency_format

    ws1['A11'] = "Ocean Freight:"
    ws1['B11'].number_format = currency_format

    ws1['A12'] = "Terminal Handling/Services:"
    ws1['B12'].number_format = currency_format

    ws1['A13'] = "Cartage & Services:"
    ws1['B13'].number_format = currency_format

    ws1['A14'] = "Other Fees:"
    ws1['B14'].number_format = currency_format

    # Total Freight Invoice
    ws1['A16'] = "Total Freight Invoice:"
    ws1['B16'] = "=SUM(B9:B14)"
    ws1['B16'].number_format = currency_format
    ws1['B16'].font = bold_font
    ws1['C16'] = "From Freight Expediters Invoice"

    # Freight Differential
    ws1['A18'] = "FREIGHT DIFFERENTIAL TO ALLOCATE:"
    ws1['B18'] = "=B16-B8"
    ws1['B18'].number_format = currency_format
    ws1['B18'].font = bold_font
    ws1['B18'].fill = yellow_fill
    ws1['C18'] = "Amount to distribute across parts"

    # Section B: Part Detail Table (Starting Row 20)
    headers = [
        "Part #", "Description", "Quantity", "Net Weight (KG)",
        "# of Skids", "Weight per Skid", "Total Part Weight",
        "Weight %", "Allocated Freight"
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws1.cell(row=20, column=col_idx)
        cell.value = header
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Data entry rows (21-70) with formulas
    for row in range(21, 71):
        # F: Weight per Skid = IF(E>0, D*C/E, 0)
        ws1[f'F{row}'] = f'=IF(E{row}>0, D{row}*C{row}/E{row}, 0)'
        ws1[f'F{row}'].number_format = '0.00'

        # G: Total Part Weight = C*D
        ws1[f'G{row}'] = f'=C{row}*D{row}'
        ws1[f'G{row}'].number_format = '0.00'

        # H: Weight % = IF($G$71>0, G/$G$71, 0)
        ws1[f'H{row}'] = f'=IF($G$71>0, G{row}/$G$71, 0)'
        ws1[f'H{row}'].number_format = percentage_format

        # I: Allocated Freight = H*$B$18
        ws1[f'I{row}'] = f'=H{row}*$B$18'
        ws1[f'I{row}'].number_format = currency_format

        # Format quantity and weight columns
        ws1[f'C{row}'].number_format = number_format
        ws1[f'D{row}'].number_format = '0.00'
        ws1[f'E{row}'].number_format = number_format

    # Row 71: Totals
    ws1['A71'] = "TOTALS:"
    ws1['A71'].font = bold_font
    ws1['C71'] = "=SUM(C21:C70)"
    ws1['C71'].font = bold_font
    ws1['C71'].number_format = number_format
    ws1['E71'] = "=SUM(E21:E70)"
    ws1['E71'].font = bold_font
    ws1['E71'].number_format = number_format
    ws1['G71'] = "=SUM(G21:G70)"
    ws1['G71'].font = bold_font
    ws1['G71'].number_format = '0.00'
    ws1['H71'] = "=SUM(H21:H70)"
    ws1['H71'].font = bold_font
    ws1['H71'].number_format = percentage_format
    ws1['I71'] = "=SUM(I21:I70)"
    ws1['I71'].font = bold_font
    ws1['I71'].number_format = currency_format

    # Row 73: Validation Check
    ws1['A73'] = "Validation Check:"
    ws1['A73'].font = bold_font
    ws1['B73'] = '=IF(ABS(I71-B18)<0.01, "✓ PASS", "✗ FAIL - Allocation Error")'
    ws1['B73'].font = bold_font

    # Row 75: Instructions
    ws1['A75'] = "INSTRUCTIONS: Enter CBP Entry Summary data in Section A. Enter Freight Expediters invoice total. Enter part details starting row 21. Freight will auto-allocate by weight percentage."
    ws1.merge_cells('A75:I75')
    ws1['A75'].alignment = Alignment(wrap_text=True)

    # Set column widths for Sheet 1
    ws1.column_dimensions['A'].width = 32
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 18
    ws1.column_dimensions['E'].width = 12
    ws1.column_dimensions['F'].width = 16
    ws1.column_dimensions['G'].width = 18
    ws1.column_dimensions['H'].width = 12
    ws1.column_dimensions['I'].width = 18

    # Freeze panes at row 3
    ws1.freeze_panes = 'A21'

    # ============================================================================
    # SHEET 2: Freight Calculator
    # ============================================================================

    # Row 1: Title
    ws2.merge_cells('A1:J1')
    ws2['A1'] = "FREIGHT COST ANALYSIS"
    ws2['A1'].font = header_font
    ws2['A1'].alignment = center_align

    # Row 2: Auto-populated import info
    ws2['A2'] = "Entry #:"
    ws2['B2'] = "='Import Data Entry'!B3"
    ws2['D2'] = "Vessel:"
    ws2['E2'] = "='Import Data Entry'!B5"
    ws2['G2'] = "Total Freight:"
    ws2['H2'] = "='Import Data Entry'!B18"
    ws2['H2'].number_format = currency_format

    # Row 4: Headers
    headers2 = [
        "Part #", "Description", "Quantity (M)", "Net Weight (KG)",
        "Allocated Freight", "Cost per M", "Cost per Unit", "Cost per KG",
        "DRM Code", "Notes"
    ]

    for col_idx, header in enumerate(headers2, start=1):
        cell = ws2.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Rows 5-54: Auto-populated from Import Data Entry
    for row in range(5, 55):
        source_row = row + 16  # Maps to rows 21-70 on Import Data Entry

        # A: Part #
        ws2[f'A{row}'] = f"='Import Data Entry'!A{source_row}"

        # B: Description
        ws2[f'B{row}'] = f"='Import Data Entry'!B{source_row}"

        # C: Quantity (M) - convert to thousands
        ws2[f'C{row}'] = f"='Import Data Entry'!C{source_row}/1000"
        ws2[f'C{row}'].number_format = '0.000'

        # D: Net Weight (KG)
        ws2[f'D{row}'] = f"='Import Data Entry'!D{source_row}"
        ws2[f'D{row}'].number_format = '0.00'

        # E: Allocated Freight
        ws2[f'E{row}'] = f"='Import Data Entry'!I{source_row}"
        ws2[f'E{row}'].number_format = currency_format

        # F: Cost per M
        ws2[f'F{row}'] = f'=IF(C{row}>0, E{row}/C{row}, 0)'
        ws2[f'F{row}'].number_format = currency_format

        # G: Cost per Unit
        ws2[f'G{row}'] = f'=IF(C{row}>0, E{row}/(C{row}*1000), 0)'
        ws2[f'G{row}'].number_format = '"$"0.000000'

        # H: Cost per KG
        ws2[f'H{row}'] = f'=IF(D{row}>0, E{row}/(D{row}*C{row}*1000), 0)'
        ws2[f'H{row}'].number_format = '"$"0.000000'

    # Row 56: Totals
    ws2['A56'] = "TOTALS:"
    ws2['A56'].font = bold_font
    ws2['E56'] = "=SUM(E5:E54)"
    ws2['E56'].font = bold_font
    ws2['E56'].number_format = currency_format

    # Set column widths for Sheet 2
    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 30
    ws2.column_dimensions['C'].width = 14
    ws2.column_dimensions['D'].width = 18
    ws2.column_dimensions['E'].width = 18
    ws2.column_dimensions['F'].width = 14
    ws2.column_dimensions['G'].width = 16
    ws2.column_dimensions['H'].width = 14
    ws2.column_dimensions['I'].width = 12
    ws2.column_dimensions['J'].width = 20

    # Freeze panes at row 3
    ws2.freeze_panes = 'A5'

    # ============================================================================
    # SHEET 3: ERP Upload
    # ============================================================================

    # Row 1: Warning
    ws3.merge_cells('A1:E1')
    ws3['A1'] = "ERP IMPORT TEMPLATE - DO NOT MODIFY HEADERS"
    ws3['A1'].font = red_font
    ws3['A1'].alignment = center_align

    # Row 2: Headers
    headers3 = [
        "Part_Number", "DRM_Code", "Freight_Cost_Per_M",
        "Freight_Cost_Per_Unit", "Entry_Number"
    ]

    for col_idx, header in enumerate(headers3, start=1):
        cell = ws3.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Rows 3-52: Auto-populated from Freight Calculator
    for row in range(3, 53):
        source_row = row + 2  # Maps to rows 5-54 on Freight Calculator

        # A: Part_Number
        ws3[f'A{row}'] = f"='Freight Calculator'!A{source_row}"

        # B: DRM_Code
        ws3[f'B{row}'] = f"='Freight Calculator'!I{source_row}"

        # C: Freight_Cost_Per_M
        ws3[f'C{row}'] = f"=ROUND('Freight Calculator'!F{source_row},4)"
        ws3[f'C{row}'].number_format = '0.0000'

        # D: Freight_Cost_Per_Unit
        ws3[f'D{row}'] = f"=ROUND('Freight Calculator'!G{source_row},6)"
        ws3[f'D{row}'].number_format = '0.000000'

        # E: Entry_Number
        ws3[f'E{row}'] = "='Import Data Entry'!B3"

    # Set column widths for Sheet 3
    ws3.column_dimensions['A'].width = 18
    ws3.column_dimensions['B'].width = 14
    ws3.column_dimensions['C'].width = 20
    ws3.column_dimensions['D'].width = 22
    ws3.column_dimensions['E'].width = 16

    # Freeze panes at row 3
    ws3.freeze_panes = 'A3'

    # ============================================================================
    # Named Ranges
    # ============================================================================

    from openpyxl.workbook.defined_name import DefinedName

    # Create named ranges using modern method
    wb.defined_names['TotalFreight'] = DefinedName('TotalFreight', attr_text="'Import Data Entry'!$B$18")
    wb.defined_names['EntryNumber'] = DefinedName('EntryNumber', attr_text="'Import Data Entry'!$B$3")
    wb.defined_names['ValidationStatus'] = DefinedName('ValidationStatus', attr_text="'Import Data Entry'!$B$73")

    # ============================================================================
    # Add Example Data
    # ============================================================================

    # Entry Summary Data
    ws1['B3'] = "2026-001-ABC"
    ws1['B4'] = datetime(2026, 1, 6)
    ws1['B5'] = "MV PACIFIC TRADER"
    ws1['B6'] = "INV-2026-001"

    # Cost breakdown
    ws1['B8'] = 5000.00  # Total Duty
    ws1['B9'] = 250.00   # Customs Entry Fee
    ws1['B10'] = 150.00  # ISF Security Filing
    ws1['B11'] = 8500.00 # Ocean Freight
    ws1['B12'] = 1200.00 # Terminal Handling
    ws1['B13'] = 850.00  # Cartage & Services
    ws1['B14'] = 300.00  # Other Fees

    # Example part data (3 parts)
    parts_data = [
        # Part #, Description, Quantity, Net Weight (KG), # of Skids
        ["PN-12345", "Widget Assembly Type A", 5000, 2.5, 10],
        ["PN-67890", "Widget Assembly Type B", 3000, 3.2, 8],
        ["PN-24680", "Widget Assembly Type C", 7500, 1.8, 15],
    ]

    for idx, part_data in enumerate(parts_data, start=21):
        ws1[f'A{idx}'] = part_data[0]  # Part #
        ws1[f'B{idx}'] = part_data[1]  # Description
        ws1[f'C{idx}'] = part_data[2]  # Quantity
        ws1[f'D{idx}'] = part_data[3]  # Net Weight
        ws1[f'E{idx}'] = part_data[4]  # # of Skids

    # Add DRM codes on Freight Calculator sheet
    ws2['I5'] = "DRM-A"
    ws2['I6'] = "DRM-B"
    ws2['I7'] = "DRM-C"

    # ============================================================================
    # Conditional Formatting for Validation
    # ============================================================================

    # Apply conditional formatting to validation cell
    from openpyxl.formatting.rule import CellIsRule

    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")

    ws1.conditional_formatting.add('B73',
        CellIsRule(operator='containsText', formula=['"PASS"'], fill=green_fill))
    ws1.conditional_formatting.add('B73',
        CellIsRule(operator='containsText', formula=['"FAIL"'], fill=red_fill))

    # ============================================================================
    # Page Setup
    # ============================================================================

    for ws in [ws1, ws2, ws3]:
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.print_options.horizontalCentered = True

    # Save the workbook
    wb.save("Die_Co_Freight_Calculator.xlsx")
    print("✓ Excel workbook 'Die_Co_Freight_Calculator.xlsx' created successfully!")
    print("\nWorkbook includes:")
    print("  • Sheet 1: Import Data Entry - Main data entry with freight allocation")
    print("  • Sheet 2: Freight Calculator - Cost analysis and calculations")
    print("  • Sheet 3: ERP Upload - Export template for ERP system")
    print("\nExample data has been added to demonstrate functionality.")
    print("\nNamed ranges created:")
    print("  • TotalFreight = 'Import Data Entry'!B18")
    print("  • EntryNumber = 'Import Data Entry'!B3")
    print("  • ValidationStatus = 'Import Data Entry'!B73")

if __name__ == "__main__":
    create_freight_calculator()
