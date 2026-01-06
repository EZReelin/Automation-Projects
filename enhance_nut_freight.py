#!/usr/bin/env python3
"""
Enhance Nut Freight Costs Excel Workbook
Creates enhanced workbook with Import Data Entry automation
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from datetime import datetime
import os

def create_base_workbook():
    """Create base workbook structure if file doesn't exist"""
    wb = Workbook()
    wb.remove(wb.active)

    # Create base sheets that would exist in original file
    ws_calc = wb.create_sheet("Shipment Calculation", 0)
    ws_hist = wb.create_sheet("Sheet1", 1)
    ws_matrix = wb.create_sheet("Sheet2", 2)

    # Setup basic Shipment Calculation structure (will be enhanced later)
    ws_calc['A1'] = "Part #"
    ws_calc['B1'] = "Net Weight(KG)"
    ws_calc['C1'] = "QTY(M)"
    ws_calc['D1'] = "Total Weight(KGS)"
    ws_calc['E1'] = "# of Skids"
    ws_calc['F1'] = "Weight %"
    ws_calc['G1'] = "Cost Allocated"
    ws_calc['H1'] = "Cost/KGS"
    ws_calc['I1'] = "Cost/M"

    # Add some basic formulas structure
    for row in range(3, 38):
        ws_calc[f'D{row}'] = f'=IF(B{row}="","",B{row}*C{row})'
        ws_calc[f'F{row}'] = f'=IF(D{row}="","",D{row}/$D$44)'
        ws_calc[f'G{row}'] = f'=IF(F{row}="","",F{row}*$B$42)'
        ws_calc[f'H{row}'] = f'=IF(D{row}="","",G{row}/D{row})'
        ws_calc[f'I{row}'] = f'=IF(C{row}="","",G{row}/C{row})'

    # Add input cells for old structure
    ws_calc['A40'] = "Totals"
    ws_calc['B42'] = "Shipment cost"
    ws_calc['D42'] = "Total # of Skids:"
    ws_calc['B44'] = "Total weight(KGS)"
    ws_calc['D44'] = "Shipment weight KGS:"
    ws_calc['D45'] = "Vessel name:"

    # Setup historical log (Sheet1)
    ws_hist['A1'] = "Part #"
    ws_hist['B1'] = "Freight Cost/M"
    ws_hist['C1'] = "Date"
    ws_hist['D1'] = "Net Weight"

    # Setup matrix reference (Sheet2)
    ws_matrix['A1'] = "Date Reference Matrix"
    ws_matrix['A2'] = "Part #"
    ws_matrix['B2'] = "Cost History"

    return wb

def enhance_workbook():
    """Create or enhance the Nut Freight Costs workbook"""

    # Check if file exists, if not create base structure
    filename = "Nut_Freight_Costs_copy.xlsx"

    if os.path.exists(filename):
        wb = load_workbook(filename)
    else:
        print("Creating base workbook structure...")
        wb = create_base_workbook()

    # Define styles
    header_font = Font(bold=True, size=14)
    bold_font = Font(bold=True)
    red_font = Font(bold=True, color="FF0000")
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    wrap_align = Alignment(wrap_text=True, vertical="top")

    currency_format = '"$"#,##0.00'
    date_format = 'MM/DD/YYYY'
    number_format = '#,##0'
    decimal2_format = '0.00'
    decimal3_format = '0.000'
    percentage_format = '0.00%'

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ============================================================================
    # ADD NEW SHEET 1: Import Data Entry (insert at beginning)
    # ============================================================================

    # Remove existing Import Data Entry if it exists
    if "Import Data Entry" in wb.sheetnames:
        del wb["Import Data Entry"]

    ws_import = wb.create_sheet("Import Data Entry", 0)

    # Section A: Source Document Inputs (Rows 1-15)
    ws_import.merge_cells('A1:F1')
    ws_import['A1'] = "DIE CO. NUT FREIGHT - IMPORT DATA ENTRY"
    ws_import['A1'].font = header_font
    ws_import['A1'].alignment = center_align

    ws_import['A3'] = "STEP 1: Enter Customs & Freight Data"
    ws_import['A3'].font = bold_font

    ws_import['A4'] = "Entry Number:"
    ws_import['B4'].fill = green_fill

    ws_import['A5'] = "Vessel Name:"
    ws_import['B5'].fill = green_fill

    ws_import['A6'] = "Entry Date:"
    ws_import['B6'].number_format = date_format
    ws_import['B6'].fill = green_fill

    ws_import['A8'] = "Total Duty (CBP Box 37):"
    ws_import['B8'].number_format = currency_format
    ws_import['B8'].fill = green_fill

    ws_import['A9'] = "Freight Invoice Total:"
    ws_import['B9'].number_format = currency_format
    ws_import['B9'].fill = green_fill

    ws_import['A10'] = "Shipment Weight (KGS):"
    ws_import['B10'].number_format = decimal2_format
    ws_import['B10'].fill = green_fill
    ws_import['C10'] = "From BOL or customs doc"

    ws_import['A11'] = "Number of Skids:"
    ws_import['B11'].number_format = number_format
    ws_import['B11'].fill = green_fill

    ws_import['A13'] = "NET FREIGHT TO ALLOCATE:"
    ws_import['B13'] = "=B9-B8"
    ws_import['B13'].number_format = currency_format
    ws_import['B13'].font = bold_font
    ws_import['B13'].fill = yellow_fill

    ws_import['A15'] = "STEP 2: Paste Part Details Below"
    ws_import['A15'].font = bold_font

    # Section B: Part Detail Table (Rows 16-50)
    headers_import = ["Part #", "Net Weight (KG)", "QTY (M)", "# of Skids", "Status"]

    for col_idx, header in enumerate(headers_import, start=1):
        cell = ws_import.cell(row=16, column=col_idx)
        cell.value = header
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Rows 17-50: Input area with formulas
    for row in range(17, 51):
        ws_import[f'A{row}'].fill = green_fill
        ws_import[f'B{row}'].number_format = decimal2_format
        ws_import[f'B{row}'].fill = green_fill
        ws_import[f'C{row}'].number_format = decimal3_format
        ws_import[f'C{row}'].fill = green_fill
        ws_import[f'D{row}'].number_format = number_format
        ws_import[f'D{row}'].fill = green_fill
        ws_import[f'E{row}'] = f'=IF(AND(A{row}<>"",B{row}<>"",C{row}<>"",D{row}<>""),"✓","")'

    # Row 52: Ready to Calculate
    ws_import['A52'] = "Ready to Calculate:"
    ws_import['A52'].font = bold_font
    ws_import['B52'] = '=IF(COUNTIF(E17:E50,"✓")>0,"YES - Go to Shipment Calculation tab","NO - Complete part details")'
    ws_import['B52'].font = bold_font

    # Instructions text box
    ws_import['A54'] = """NEW PROCESS:
1. Enter customs entry number, vessel, date
2. Enter total duty from CBP Box 37
3. Enter freight invoice total (ALL charges)
4. Enter shipment weight (KGS) and skid count
5. Paste or type part details (Part #, Net Weight, QTY, Skids)
6. Go to Shipment Calculation tab - review auto-calculated costs
7. Go to ERP Upload tab - copy table and paste into ERP system
8. Record in Historical Log for future reference"""
    ws_import.merge_cells('A54:F62')
    ws_import['A54'].alignment = wrap_align
    ws_import['A54'].border = thin_border

    # Set column widths
    ws_import.column_dimensions['A'].width = 32
    ws_import.column_dimensions['B'].width = 18
    ws_import.column_dimensions['C'].width = 15
    ws_import.column_dimensions['D'].width = 15
    ws_import.column_dimensions['E'].width = 10
    ws_import.column_dimensions['F'].width = 25

    # Freeze panes at row 17
    ws_import.freeze_panes = 'A17'

    # ============================================================================
    # MODIFY EXISTING SHEET: Shipment Calculation
    # ============================================================================

    ws_calc = wb["Shipment Calculation"]

    # Insert 2 rows at top
    ws_calc.insert_rows(1, 2)

    # Add warning header
    ws_calc.merge_cells('A1:I1')
    ws_calc['A1'] = "AUTO-POPULATED FROM IMPORT DATA ENTRY - DO NOT EDIT"
    ws_calc['A1'].font = red_font
    ws_calc['A1'].alignment = center_align

    # Update headers (now in row 3)
    headers_calc = ["Part #", "Net Weight(KG)", "QTY(M)", "Total Weight(KGS)",
                    "# of Skids", "Weight %", "Cost Allocated", "Cost/KGS", "Cost/M"]

    for col_idx, header in enumerate(headers_calc, start=1):
        cell = ws_calc.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Rows 5-39: Auto-populated formulas (shifted by 2 from original rows 3-37)
    for row in range(5, 40):
        idx = row - 4

        # A: Part # from Import Data Entry
        ws_calc[f'A{row}'] = f'=IFERROR(INDEX(\'Import Data Entry\'!$A$17:$A$50,{idx}),"")'

        # B: Net Weight from Import Data Entry
        ws_calc[f'B{row}'] = f'=IFERROR(INDEX(\'Import Data Entry\'!$B$17:$B$50,{idx}),"")'
        ws_calc[f'B{row}'].number_format = decimal2_format

        # C: QTY(M) from Import Data Entry
        ws_calc[f'C{row}'] = f'=IFERROR(INDEX(\'Import Data Entry\'!$C$17:$C$50,{idx}),"")'
        ws_calc[f'C{row}'].number_format = decimal3_format

        # D: Total Weight calculation (Net Weight is KG per M, so just multiply by QTY in M)
        ws_calc[f'D{row}'] = f'=IF(B{row}="","",B{row}*C{row})'
        ws_calc[f'D{row}'].number_format = decimal2_format

        # E: # of Skids from Import Data Entry
        ws_calc[f'E{row}'] = f'=IFERROR(INDEX(\'Import Data Entry\'!$D$17:$D$50,{idx}),"")'
        ws_calc[f'E{row}'].number_format = number_format

        # F: Weight % (divide by sum of all part weights, not shipment weight)
        ws_calc[f'F{row}'] = f'=IF(D{row}="","",D{row}/$D$41)'
        ws_calc[f'F{row}'].number_format = percentage_format

        # G: Cost Allocated
        ws_calc[f'G{row}'] = f'=IF(F{row}="","",F{row}*$B$44)'
        ws_calc[f'G{row}'].number_format = currency_format

        # H: Cost/KGS
        ws_calc[f'H{row}'] = f'=IF(D{row}="","",G{row}/D{row})'
        ws_calc[f'H{row}'].number_format = currency_format

        # I: Cost/M
        ws_calc[f'I{row}'] = f'=IF(C{row}="","",G{row}/C{row})'
        ws_calc[f'I{row}'].number_format = currency_format

    # Totals row (now row 41, was 40 before shift)
    ws_calc['A41'] = "Totals"
    ws_calc['A41'].font = bold_font
    ws_calc['C41'] = "=SUM(C5:C39)"
    ws_calc['C41'].font = bold_font
    ws_calc['C41'].number_format = decimal3_format
    ws_calc['D41'] = "=SUM(D5:D39)"
    ws_calc['D41'].font = bold_font
    ws_calc['D41'].number_format = decimal2_format
    ws_calc['E41'] = "=SUM(E5:E39)"
    ws_calc['E41'].font = bold_font
    ws_calc['E41'].number_format = number_format
    ws_calc['G41'] = "=SUM(G5:G39)"
    ws_calc['G41'].font = bold_font
    ws_calc['G41'].number_format = currency_format

    # Summary section (shifted by 2)
    ws_calc['A44'] = "Shipment cost"
    ws_calc['B44'] = "='Import Data Entry'!B13"
    ws_calc['B44'].number_format = currency_format
    ws_calc['B44'].font = bold_font
    ws_calc['B44'].fill = yellow_fill

    ws_calc['C44'] = "Total # of Skids:"
    ws_calc['D44'] = "='Import Data Entry'!B11"
    ws_calc['D44'].number_format = number_format
    ws_calc['D44'].font = bold_font

    ws_calc['A46'] = "Total weight(KGS)"
    ws_calc['A46'].font = bold_font
    ws_calc['C46'] = "Shipment weight KGS:"
    ws_calc['D46'] = "='Import Data Entry'!B10"
    ws_calc['D46'].number_format = decimal2_format
    ws_calc['D46'].font = bold_font
    ws_calc['D46'].fill = yellow_fill

    ws_calc['C47'] = "Vessel name:"
    ws_calc['D47'] = "='Import Data Entry'!B5"
    ws_calc['D47'].font = bold_font

    ws_calc['C48'] = "Entry Date:"
    ws_calc['D48'] = "='Import Data Entry'!B6"
    ws_calc['D48'].number_format = date_format
    ws_calc['D48'].font = bold_font

    # Instructions
    ws_calc['A50'] = "AUTO-CALCULATED: All data pulls from Import Data Entry sheet. Review calculations and copy Cost/M values to Historical Log when complete."
    ws_calc.merge_cells('A50:I52')
    ws_calc['A50'].alignment = wrap_align
    ws_calc['A50'].border = thin_border
    ws_calc['A50'].fill = header_fill

    # Set column widths
    ws_calc.column_dimensions['A'].width = 15
    ws_calc.column_dimensions['B'].width = 16
    ws_calc.column_dimensions['C'].width = 12
    ws_calc.column_dimensions['D'].width = 18
    ws_calc.column_dimensions['E'].width = 12
    ws_calc.column_dimensions['F'].width = 12
    ws_calc.column_dimensions['G'].width = 16
    ws_calc.column_dimensions['H'].width = 14
    ws_calc.column_dimensions['I'].width = 14

    # Freeze panes at row 4
    ws_calc.freeze_panes = 'A5'

    # ============================================================================
    # ADD NEW SHEET 3: ERP Upload Template
    # ============================================================================

    # Remove existing ERP Upload if it exists
    if "ERP Upload Template" in wb.sheetnames:
        del wb["ERP Upload Template"]

    ws_erp = wb.create_sheet("ERP Upload Template", 3)

    # Header
    ws_erp.merge_cells('A1:E1')
    ws_erp['A1'] = "ERP IMPORT - COPY TO CLIPBOARD"
    ws_erp['A1'].font = header_font
    ws_erp['A1'].alignment = center_align

    # Row 2: Headers
    headers_erp = ["Part_Number", "DRM_Code", "Freight_Cost_Per_M", "Entry_Date", "Vessel"]

    for col_idx, header in enumerate(headers_erp, start=1):
        cell = ws_erp.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Rows 3-37: Auto-populated
    for row in range(3, 38):
        source_row = row + 2  # Maps to rows 5-39 on Shipment Calculation

        # A: Part_Number
        ws_erp[f'A{row}'] = f"='Shipment Calculation'!A{source_row}"

        # B: DRM_Code (manual input)
        ws_erp[f'B{row}'].fill = green_fill

        # C: Freight_Cost_Per_M
        ws_erp[f'C{row}'] = f"='Shipment Calculation'!I{source_row}"
        ws_erp[f'C{row}'].number_format = '0.0000'

        # D: Entry_Date
        ws_erp[f'D{row}'] = "='Import Data Entry'!B6"
        ws_erp[f'D{row}'].number_format = date_format

        # E: Vessel
        ws_erp[f'E{row}'] = "='Import Data Entry'!B5"

    # Set column widths
    ws_erp.column_dimensions['A'].width = 18
    ws_erp.column_dimensions['B'].width = 14
    ws_erp.column_dimensions['C'].width = 20
    ws_erp.column_dimensions['D'].width = 14
    ws_erp.column_dimensions['E'].width = 25

    # Freeze panes at row 3
    ws_erp.freeze_panes = 'A3'

    # Enable autofilter on headers
    ws_erp.auto_filter.ref = 'A2:E37'

    # ============================================================================
    # MODIFY EXISTING SHEET: Sheet1 (Historical Log)
    # ============================================================================

    ws_hist = wb["Sheet1"]

    # Rename sheet
    ws_hist.title = "Historical Log"

    # Format headers if not already formatted
    for col in range(1, 5):
        cell = ws_hist.cell(row=1, column=col)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Add instruction in row 2
    ws_hist.insert_rows(2)
    ws_hist['A2'] = "→ To add current shipment data, copy Cost/M values from Shipment Calculation tab and paste below"
    ws_hist.merge_cells('A2:D2')
    ws_hist['A2'].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    ws_hist['A2'].alignment = wrap_align

    # Set column widths
    ws_hist.column_dimensions['A'].width = 18
    ws_hist.column_dimensions['B'].width = 18
    ws_hist.column_dimensions['C'].width = 14
    ws_hist.column_dimensions['D'].width = 16

    # ============================================================================
    # Named Ranges
    # ============================================================================

    wb.defined_names['NetFreight'] = DefinedName('NetFreight', attr_text="'Import Data Entry'!$B$13")
    wb.defined_names['ShipmentDate'] = DefinedName('ShipmentDate', attr_text="'Import Data Entry'!$B$6")
    wb.defined_names['VesselName'] = DefinedName('VesselName', attr_text="'Import Data Entry'!$B$5")
    wb.defined_names['TotalSkids'] = DefinedName('TotalSkids', attr_text="'Import Data Entry'!$B$11")

    # ============================================================================
    # Add Real Data from Ever Eagle Shipment
    # ============================================================================

    # Data from CBP Entry Summary and Freight Invoice
    ws_import['B4'] = "AVZ-0154383-3"  # Entry Number from CBP
    ws_import['B5'] = "EVER EAGLE 192E"  # Vessel from freight invoice
    ws_import['B6'] = datetime(2025, 12, 17)  # Entry Date from CBP
    ws_import['B8'] = 10991.50  # Total Duty from CBP Box 37
    ws_import['B9'] = 15382.36  # Freight Invoice Total
    ws_import['B10'] = 7821.00  # Shipment Weight (KGS) from freight invoice
    ws_import['B11'] = 13  # Number of Skids (total from parts)

    # Real parts data from Ever Eagle shipment
    # Data matches current spreadsheet: Part #, Net Weight (KG), QTY (M), # of Skids
    real_parts = [
        ["DCS-621", 7.25, 243.0, 2],
        ["DCS-623", 6.90, 252.0, 2],
        ["DCS-672", 15.40, 264.6, 3],
        ["DCS-929", 7.94, 1020.6, 6],
    ]

    for idx, part in enumerate(real_parts, start=17):
        ws_import[f'A{idx}'] = part[0]  # Part #
        ws_import[f'B{idx}'] = part[1]  # Net Weight (KG)
        ws_import[f'C{idx}'] = part[2]  # QTY (M)
        ws_import[f'D{idx}'] = part[3]  # # of Skids

    # Add real DRM codes from internal form
    drm_codes = {
        "DCS-621": "18735",
        "DCS-623": "18732",
        "DCS-672": "18736",
        "DCS-929": "18734",
    }

    # Map DRM codes to ERP sheet rows (row 3 maps to DCS-621, etc.)
    for i, (part_num, drm) in enumerate(drm_codes.items(), start=3):
        ws_erp[f'B{i}'] = drm

    # ============================================================================
    # Page Setup
    # ============================================================================

    for sheet_name in ["Import Data Entry", "Shipment Calculation", "ERP Upload Template"]:
        ws = wb[sheet_name]
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.print_options.horizontalCentered = True

    # ============================================================================
    # Conditional Formatting
    # ============================================================================

    from openpyxl.formatting.rule import CellIsRule

    green_conditional = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_conditional = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Status column conditional formatting
    ws_import.conditional_formatting.add('E17:E50',
        CellIsRule(operator='equal', formula=['"✓"'], fill=green_conditional))

    # Ready to calculate conditional formatting
    ws_import.conditional_formatting.add('B52',
        CellIsRule(operator='containsText', formula=['"YES"'], fill=green_conditional))
    ws_import.conditional_formatting.add('B52',
        CellIsRule(operator='containsText', formula=['"NO"'], fill=red_conditional))

    # Save the workbook
    wb.save("Nut_Freight_Costs_copy.xlsx")
    print("✓ Enhanced workbook 'Nut_Freight_Costs_copy.xlsx' created successfully!")
    print("\nEnhancements added:")
    print("  • NEW: Import Data Entry sheet - Streamlined data entry with validation")
    print("  • MODIFIED: Shipment Calculation - Auto-populates from Import Data Entry")
    print("  • NEW: ERP Upload Template - Ready for system import")
    print("  • MODIFIED: Historical Log - Enhanced with instructions")
    print("\nNamed ranges created:")
    print("  • NetFreight = 'Import Data Entry'!B13")
    print("  • ShipmentDate = 'Import Data Entry'!B6")
    print("  • VesselName = 'Import Data Entry'!B5")
    print("  • TotalSkids = 'Import Data Entry'!B11")
    print("\nExample shipment data has been added to demonstrate workflow.")

if __name__ == "__main__":
    enhance_workbook()
